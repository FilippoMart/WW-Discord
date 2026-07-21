import openpyxl
import sqlite3
import os

SRC = r"C:\Users\filip\Documents\_Filippo stuff\Giochi\ww\WWDiscordtemplate.xlsx"
OUT_DIR = r"C:\Users\filip\Documents\_Filippo stuff\Giochi\ww"
DB_PATH = os.path.join(OUT_DIR, "WWDiscordtemplate.sqlite")
SQL_PATH = os.path.join(OUT_DIR, "WWDiscordtemplate.sql")

if os.path.exists(DB_PATH):
    os.remove(DB_PATH)

wb = openpyxl.load_workbook(SRC, data_only=True)

conn = sqlite3.connect(DB_PATH)
cur = conn.cursor()
cur.executescript("""
CREATE TABLE players (
    nick TEXT PRIMARY KEY,
    nome TEXT NOT NULL
);

CREATE TABLE roles (
    ruolo TEXT PRIMARY KEY,
    aura TEXT,       -- P = Pura, O = Oscura, alignment as seen by a Veggente
    misticismo TEXT, -- M = Mistico, NM = Non Mistico
    fazione TEXT,
    ombra TEXT,      -- 'Ombra' o 'Non Ombra'
    espansione TEXT,
    set_lune TEXT    -- '1 luna', '2 lune', '3 lune', o vuoto
);

CREATE TABLE games (
    play TEXT PRIMARY KEY
);

CREATE TABLE game_notes (
    play TEXT PRIMARY KEY REFERENCES games(play),
    notes TEXT
);

CREATE TABLE game_wins (
    play TEXT PRIMARY KEY REFERENCES games(play),
    win TEXT,       -- winning faction(s)
    win_ruoli TEXT  -- winning starting roles; filled in manually after each game, often blank
);

CREATE TABLE game_ruoli_possibili (
    play TEXT PRIMARY KEY REFERENCES games(play),
    ruoli_possibili TEXT  -- description of the role set/ruleset used for that game
);

CREATE TABLE assignments (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    play TEXT NOT NULL REFERENCES games(play),
    nick TEXT NOT NULL REFERENCES players(nick),
    ruolo TEXT,
    UNIQUE(play, nick)
);

CREATE TABLE fazione_overrides (
    play TEXT NOT NULL REFERENCES games(play),
    nick TEXT NOT NULL REFERENCES players(nick),
    fazione_finale TEXT,  -- condizioni vittoria finali, sostituisce r.fazione per questo assignment (es. vampirizzato, gildato, romeo, branco svegliato)
    PRIMARY KEY(play, nick)
);
""")

# --- players ---
ws = wb["Giocatori"]
rows = list(ws.iter_rows(values_only=True))[1:]  # skip header
players = [(r[0], r[1]) for r in rows if r[0] is not None]
cur.executemany("INSERT INTO players (nick, nome) VALUES (?, ?)", players)

# --- roles ---
ws = wb["Ruoli"]
rows = list(ws.iter_rows(values_only=True))[1:]
roles = [(r[0], r[1], r[2], r[3], r[4], r[5], r[6]) for r in rows if r[0] is not None]
cur.executemany("INSERT INTO roles (ruolo, aura, misticismo, fazione, ombra, espansione, set_lune) VALUES (?, ?, ?, ?, ?, ?, ?)", roles)

# --- games + assignments (wide -> long) from GPR ---
ws = wb["GPR"]
rows = list(ws.iter_rows(values_only=True))
header = rows[0]
play_cols = header[2:]  # columns from index 2 onward are play ids

games = [(p,) for p in play_cols if p is not None]
cur.executemany("INSERT OR IGNORE INTO games (play) VALUES (?)", games)

assignments = []
for r in rows[1:]:
    nick = r[0]
    if nick is None:
        continue
    for col_idx, play in enumerate(play_cols, start=2):
        if play is None:
            continue
        ruolo = r[col_idx] if col_idx < len(r) else None
        if ruolo is not None:
            assignments.append((play, nick, ruolo))
cur.executemany("INSERT INTO assignments (play, nick, ruolo) VALUES (?, ?, ?)", assignments)

# --- game_notes ---
ws = wb["PNotes"]
rows = list(ws.iter_rows(values_only=True))[1:]
notes = [(r[0], r[1]) for r in rows if r[0] is not None]
for play, _ in notes:
    cur.execute("INSERT OR IGNORE INTO games (play) VALUES (?)", (play,))
cur.executemany("INSERT INTO game_notes (play, notes) VALUES (?, ?)", notes)

# --- game_wins ---
ws = wb["PWin"]
rows = list(ws.iter_rows(values_only=True))[1:]
wins = [(r[0], r[1], r[2]) for r in rows if r[0] is not None]
for play, _, _ in wins:
    cur.execute("INSERT OR IGNORE INTO games (play) VALUES (?)", (play,))
cur.executemany("INSERT INTO game_wins (play, win, win_ruoli) VALUES (?, ?, ?)", wins)

# --- game_ruoli_possibili ---
ws = wb["PRuoliPoss"]
rows = list(ws.iter_rows(values_only=True))[1:]
ruoli_possibili = [(r[0], r[1]) for r in rows if r[0] is not None]
for play, _ in ruoli_possibili:
    cur.execute("INSERT OR IGNORE INTO games (play) VALUES (?)", (play,))
cur.executemany("INSERT INTO game_ruoli_possibili (play, ruoli_possibili) VALUES (?, ?)", ruoli_possibili)

# --- fazione_overrides ---
ws = wb["POverrides"]
rows = list(ws.iter_rows(values_only=True))[1:]
overrides = [(r[0], r[1], r[2]) for r in rows if r[0] is not None]
cur.executemany("INSERT INTO fazione_overrides (play, nick, fazione_finale) VALUES (?, ?, ?)", overrides)

conn.commit()

# --- dump to .sql file ---
with open(SQL_PATH, "w", encoding="utf-8") as f:
    for line in conn.iterdump():
        f.write(f"{line}\n")

conn.close()

print("players:", len(players))
print("roles:", len(roles))
print("games:", len(set(p for p, in games)))
print("assignments:", len(assignments))
print("notes:", len(notes))
print("wins:", len(wins))
print("ruoli_possibili:", len(ruoli_possibili))
print("fazione_overrides:", len(overrides))
print("DB written to:", DB_PATH)
print("SQL written to:", SQL_PATH)
